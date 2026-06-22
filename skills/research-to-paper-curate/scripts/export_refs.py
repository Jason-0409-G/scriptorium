#!/usr/bin/env python3
"""Export a verified paper list to a reference library: RIS + BibTeX + by-category Excel.

Why RIS is the default: a single `.ris` file imports natively into BOTH Zotero and EndNote,
so we avoid tool-specific APIs and credential setup. `.bib` is added for LaTeX users. The
Excel is for the human to read and prune by category before importing.

Run this only on papers that passed DOI verification + the adversarial review gate. It prefers
`verified_doi` (set by verify_doi.py) over the raw `doi`, so a corrected DOI flows through.

Stdlib for RIS/BibTeX (plain text). Excel uses openpyxl if available; otherwise a .csv is written
with the same content so the step never hard-fails on a missing dependency.

Input  : verified papers (.json / .tsv / .csv). Fields: title, authors, year, journal,
         doi or verified_doi, category, abstract, doi_status.
Output : <outdir>/library.ris, library.bib, and library.xlsx (or library.csv fallback).

Usage: python export_refs.py <verified.json|tsv|csv> <outdir>
"""
import sys, os, json, csv, re, html, argparse

STATUS_FILL = {"verified": "C6EFCE", "candidate": "FFEB9C", "mismatch": "FFC7CE",
               "dead": "FFC7CE", "no_doi": "F2F2F2"}   # green / yellow / red / red / grey


def clean_text(s):
    """Strip XML/JATS tags, unescape entities, collapse whitespace to one line."""
    if not s:
        return ""
    if isinstance(s, list):
        s = " ".join(str(x) for x in s)
    s = re.sub(r"<[^>]+>", "", str(s))
    s = html.unescape(s)
    return " ".join(s.split())


def read_rows(path):
    if path.lower().endswith(".json"):
        d = json.load(open(path, encoding="utf-8"))
        return d if isinstance(d, list) else d.get("papers", [])
    delim = "\t" if path.lower().endswith(".tsv") else ","
    with open(path, encoding="utf-8-sig", newline="") as f:
        return [dict(r) for r in csv.DictReader(f, delimiter=delim)]


def author_list(row):
    """Return a list of authors from a string ('A; B' / 'A and B' / 'A, B') or an actual list."""
    a = row.get("authors", "")
    if isinstance(a, list):
        return [x for x in a if x]
    a = (a or "").strip()
    if not a:
        return []
    if ";" in a:
        return [x.strip() for x in a.split(";") if x.strip()]
    if " and " in a:
        return [x.strip() for x in a.split(" and ") if x.strip()]
    return [a]                       # keep "Last, First" style intact when no clear separator


def best_doi(row):
    return (row.get("verified_doi") or row.get("doi") or "").strip()


def surname(author):
    a = author.strip()
    if "," in a:
        return a.split(",")[0].strip()
    return a.split()[-1] if a.split() else "anon"


def bib_key(row, used):
    au = author_list(row)
    k = re.sub(r"[^A-Za-z0-9]", "", surname(au[0]) if au else "anon") or "ref"
    k = f"{k}{row.get('year','')}"
    n, base = 0, k
    while k in used:
        n += 1
        suf, x = "", n
        while x:
            x, r = divmod(x - 1, 26)
            suf = chr(97 + r) + suf
        k = f"{base}{suf}"
    used.add(k)
    return k


def write_ris(rows, path):
    """RIS: imports into Zotero AND EndNote."""
    out = []
    for r in rows:
        out.append("TY  - JOUR")
        for au in author_list(r):
            if "," not in au:
                toks = au.split()
                if len(toks) > 1:
                    au = f"{toks[-1]}, {' '.join(toks[:-1])}"
            out.append(f"AU  - {au}")
        if r.get("title"):   out.append(f"TI  - {clean_text(r['title'])}")
        if r.get("year"):    out.append(f"PY  - {r['year']}")
        if r.get("journal"): out.append(f"JO  - {clean_text(r['journal'])}")
        if best_doi(r):      out.append(f"DO  - {best_doi(r)}")
        if r.get("abstract"):out.append(f"AB  - {clean_text(r['abstract'])}")
        if r.get("category"):out.append(f"KW  - {r['category']}")
        out.append("ER  - ")
        out.append("")
    open(path, "w", encoding="utf-8").write("\n".join(out))


def bibtex_escape(s):
    return (s or "").replace("&", r"\&").replace("%", r"\%").replace("_", r"\_")


def write_bib(rows, path):
    used, out = set(), []
    for r in rows:
        key = bib_key(r, used)
        au = " and ".join(author_list(r))
        out.append(f"@article{{{key},")
        out.append(f"  title = {{{bibtex_escape(clean_text(r.get('title','')))}}},")
        if au:                out.append(f"  author = {{{bibtex_escape(au)}}},")
        if r.get("journal"):  out.append(f"  journal = {{{bibtex_escape(clean_text(r['journal']))}}},")
        if r.get("year"):     out.append(f"  year = {{{r['year']}}},")
        if best_doi(r):       out.append(f"  doi = {{{best_doi(r)}}},")
        out.append("}\n")
    open(path, "w", encoding="utf-8").write("\n".join(out))


COLS = ["category", "title", "authors", "year", "journal", "verified_doi", "doi_status"]


def cell(r, c):
    if c == "verified_doi":
        return best_doi(r)
    v = r.get(c, "")
    return "; ".join(v) if isinstance(v, list) else (v or "")


def write_excel(rows, path):
    """By-category, color-coded Excel via openpyxl; CSV fallback if openpyxl is absent."""
    rows = sorted(rows, key=lambda r: (str(r.get("category", "zzz")), str(r.get("year", ""))))
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        csvp = os.path.splitext(path)[0] + ".csv"
        with open(csvp, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f); w.writerow(COLS)
            for r in rows:
                w.writerow([cell(r, c) for c in COLS])
        print(f"[export] openpyxl 缺失 → 改写 CSV: {csvp}（装 openpyxl 可得着色 Excel）")
        return
    wb = Workbook(); ws = wb.active; ws.title = "Literature"
    hdr_fill = PatternFill("solid", fgColor="2E74B7")
    for j, c in enumerate(COLS, 1):
        cl = ws.cell(1, j, c); cl.font = Font(bold=True, color="FFFFFF")
        cl.fill = hdr_fill; cl.alignment = Alignment(horizontal="center")
    for i, r in enumerate(rows, 2):
        for j, c in enumerate(COLS, 1):
            ws.cell(i, j, cell(r, c)).alignment = Alignment(vertical="top", wrap_text=(c in ("title",)))
        st = r.get("doi_status", "")
        if st in STATUS_FILL:                       # color the status cell by verification result
            ws.cell(i, len(COLS)).fill = PatternFill("solid", fgColor=STATUS_FILL[st])
    widths = {"category": 16, "title": 60, "authors": 30, "year": 6, "journal": 26, "verified_doi": 30, "doi_status": 12}
    for j, c in enumerate(COLS, 1):
        ws.column_dimensions[ws.cell(1, j).column_letter].width = widths.get(c, 18)
    ws.freeze_panes = "A2"
    wb.save(path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("infile"); ap.add_argument("outdir")
    a = ap.parse_args()
    rows = read_rows(a.infile)
    if not rows:
        sys.exit("[export] 输入为空 / empty input")
    os.makedirs(a.outdir, exist_ok=True)
    write_ris(rows, os.path.join(a.outdir, "library.ris"))
    write_bib(rows, os.path.join(a.outdir, "library.bib"))
    write_excel(rows, os.path.join(a.outdir, "library.xlsx"))
    cats = {}
    for r in rows:
        cats[r.get("category", "(未分类)")] = cats.get(r.get("category", "(未分类)"), 0) + 1
    print(f"[export] {len(rows)} 篇 → {a.outdir}/  (library.ris / library.bib / library.xlsx)")
    print(f"[export] 分类: {cats}")
    print("[export] 把 library.ris 导入 Zotero 或 EndNote 即可（两者都原生支持 RIS）")


if __name__ == "__main__":
    main()
